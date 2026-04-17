"""
知识库问答测试驱动器
"""

import asyncio
import time
import json
import logging
from pathlib import Path
from dataclasses import dataclass, field
from typing import Optional, List, Dict, Any
from datetime import datetime

from ..core.config import Config, QATestConfig, EndpointConfig
from ..core.auth import AuthManager
from ..core.client import HttpClient

# 配置日志
logger = logging.getLogger(__name__)
logging.getLogger("httpx").setLevel(logging.WARNING)

# no-answer 关键词列表（与前端 NO_ANSWER_PATTERNS_ZH 保持一致）
NO_ANSWER_PATTERNS = [
    "未能找到相關資料", "未能找到相关资料",
    "沒有找到相關", "没有找到相关",
    "沒有相關資料", "没有相关资料",
    "提供的參考資料中沒有", "提供的参考资料中没有",
    "參考資料中沒有關於", "参考资料中没有关于",
    "無法回答", "无法回答",
    "沒有相關信息", "没有相关信息",
    "找不到相關", "找不到相关"
]

# 信封模板
ENVELOPE_TEMPLATE_ZH = """敬啟者：

抱歉，未能找到相關資料

香港海關"""

ENVELOPE_TEMPLATE_EN = """Dear Sir/Madam,

Sorry, no relevant information was found.

Hong Kong Customs"""

# 配置流式查询请求日志文件
def setup_request_logger(log_dir: str = "logs"):
    """配置流式查询请求日志记录器"""
    request_logger = logging.getLogger("qa_request")
    request_logger.setLevel(logging.INFO)
    request_logger.propagate = False  # 避免重复打印

    # 清除已有的 handlers
    request_logger.handlers.clear()

    # 创建日志目录
    log_path = Path(log_dir)
    log_path.mkdir(parents=True, exist_ok=True)

    # 添加文件处理器
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    file_handler = logging.FileHandler(
        log_path / f"qa_request_{timestamp}.log",
        encoding="utf-8"
    )
    file_handler.setLevel(logging.INFO)
    file_handler.setFormatter(logging.Formatter("%(asctime)s - %(message)s"))

    request_logger.addHandler(file_handler)
    return request_logger

# 全局请求日志记录器
request_logger = logging.getLogger("qa_request")


@dataclass
class SessionInfo:
    """会话信息"""
    session_id: str
    title: str
    pinned: bool = False
    created_at: Optional[str] = None
    updated_at: Optional[str] = None


@dataclass
class MessageInfo:
    """消息信息"""
    message_id: str
    session_id: str
    role: str
    content: str
    created_at: Optional[str] = None


@dataclass
class FinalAnswer:
    """最终答复结构（与前端 UI 对齐）"""
    answer_type: str = "empty"  # "intent_email" | "no_answer" | "rag_answer" | "error" | "empty"
    display_body: str = ""       # UI 实际渲染的完整 markdown
    plain_text: str = ""         # 剥掉信封头尾后的纯文字
    lang: str = "zh-hk"
    intent_quote: Dict[str, str] = field(default_factory=dict)  # {"intent_id": "", "intent_name": ""}
    citations: List[dict] = field(default_factory=list)
    raw_answer: str = ""         # 原始 LLM 回答（未处理前）


@dataclass
class QAResult:
    """单次问答结果"""
    question: str = ""
    success: bool = False
    answer: Optional[str] = None
    answer_type: str = "empty"  # "intent_email" | "no_answer" | "rag_answer" | "error" | "empty"
    response_time: float = 0.0
    citations: List[dict] = field(default_factory=list)
    error: Optional[str] = None
    metadata: Dict[str, Any] = field(default_factory=dict)
    session_id: Optional[str] = None
    message_id: Optional[str] = None
    request_details: List[dict] = field(default_factory=list)  # 记录每个接口的请求和响应
    question_id: Optional[str] = None  # 问题编号（从测试集第一列读取）


@dataclass
class BatchQAResult:
    """批量问答结果"""
    total: int = 0
    success: int = 0
    failed: int = 0
    results: List[QAResult] = field(default_factory=list)
    start_time: datetime = None
    end_time: datetime = None

    @property
    def success_rate(self) -> float:
        """成功率"""
        if self.total == 0:
            return 0.0
        return self.success / self.total

    @property
    def avg_response_time(self) -> float:
        """平均响应时间"""
        successful = [r for r in self.results if r.success]
        if not successful:
            return 0.0
        return sum(r.response_time for r in successful) / len(successful)

    @property
    def p95_response_time(self) -> float:
        """P95 响应时间"""
        successful = sorted([r.response_time for r in self.results if r.success])
        if not successful:
            return 0.0
        index = int(len(successful) * 0.95)
        return successful[min(index, len(successful) - 1)]

    @property
    def p99_response_time(self) -> float:
        """P99 响应时间"""
        successful = sorted([r.response_time for r in self.results if r.success])
        if not successful:
            return 0.0
        index = int(len(successful) * 0.99)
        return successful[min(index, len(successful) - 1)]

    @property
    def duration(self) -> float:
        """总耗时（秒）"""
        if self.start_time and self.end_time:
            return (self.end_time - self.start_time).total_seconds()
        return 0.0


class QADriver:
    """问答测试驱动器 - 执行知识库问答测试"""

    def __init__(self, config: Config, auth_manager: AuthManager):
        """
        初始化驱动器

        Args:
            config: 配置对象
            auth_manager: 认证管理器
        """
        self.config = config
        self.auth_manager = auth_manager

    def _get_debug_sse_log(self) -> bool:
        """获取 debug_sse_log 配置"""
        return getattr(self.config, 'debug_sse_log', False)

    def _get_endpoint(self, name: str) -> Optional[EndpointConfig]:
        """获取指定名称的端点配置"""
        return self.config.get_endpoint(name)

    def _get_chat_mode(self) -> str:
        """根据 rag_create_message 配置返回 chat_mode 值"""
        endpoint = self._get_endpoint("rag_create_message")
        if endpoint and endpoint.body:
            return endpoint.body.get("chat_mode", "flex")
        return "flex"

    def _get_chat_mode_header(self) -> str:
        """根据 rag_create_message 的 chat_mode 配置返回对应的表头"""
        endpoint = self._get_endpoint("rag_create_message")
        if endpoint and endpoint.body:
            chat_mode = endpoint.body.get("chat_mode", "flex")
            if chat_mode == "pipeline":
                return "Fixed Pipeline"
            else:
                return "Flexible Skills"
        return "Flexible Skills"

    async def get_document_detail(self, document_id: str) -> Optional[Dict[str, Any]]:
        """
        获取文档详情（用于查询文档名称）

        Args:
            document_id: 文档 ID

        Returns:
            文档详情字典，包含 doc_name 等字段
        """
        client = None
        try:
            auth_header = await self.auth_manager.get_auth_header()
            endpoint = self._get_endpoint("document_detail")
            if endpoint is None:
                logger.warning("未找到端点配置：document_detail")
                return None

            client = HttpClient(base_url=self.config.get_base_url(endpoint.base))
            path = self._format_path(endpoint.path, document_id=document_id)

            response = await client.request(
                method=endpoint.method,
                path=path,
                headers=auth_header,
            )

            if response.status_code != 200:
                logger.debug(f"获取文档详情失败：{document_id} - {response.status_code}")
                return None

            return response.json()

        except Exception as e:
            logger.debug(f"获取文档详情异常：{document_id} - {e}")
            return None

        finally:
            if client:
                await client.close()

    def _format_path(self, path: str, **kwargs) -> str:
        """格式化路径中的占位符"""
        for key, value in kwargs.items():
            path = path.replace(f"{{{key}}}", str(value))
        return path

    def _build_body(self, endpoint: EndpointConfig, **kwargs) -> dict:
        """根据端点配置构建请求体"""
        body = {}
        if endpoint.body:
            for key, value in endpoint.body.items():
                if isinstance(value, str):
                    # 替换占位符
                    body[key] = value.format(**kwargs)
                else:
                    body[key] = value
        return body

    def _add_request_detail(
        self,
        result: QAResult,
        method_name: str,
        method: str,
        path: str,
        request_body: dict,
        response_data: Any,
        elapsed: float,
        status_code: int,
    ):
        """添加请求详情到指定结果"""
        result.request_details.append({
            "api": method_name,
            "method": method,
            "path": path,
            "request_body": request_body,
            "response_data": response_data if isinstance(response_data, (dict, list)) else {"raw": str(response_data)},
            "status_code": status_code,
            "elapsed": f"{elapsed:.3f}s",
        })

    async def create_session(self, title: str = "New Chat", result: QAResult = None) -> SessionInfo:
        """
        创建新会话

        Args:
            title: 会话标题
            result: 用于记录请求详情的结果对象

        Returns:
            SessionInfo 会话信息
        """
        client = None
        start_time = time.time()
        try:
            auth_header = await self.auth_manager.get_auth_header()
            endpoint = self._get_endpoint("rag_create_session")
            if endpoint is None:
                raise Exception("未找到端点配置：rag_create_session，请检查 config/endpoints.yaml")

            client = HttpClient(base_url=self.config.get_base_url(endpoint.base))
            body = self._build_body(endpoint, title=title)

            response = await client.request(
                method=endpoint.method,
                path=endpoint.path,
                headers=auth_header,
                json=body,
            )

            if response.status_code not in [200, 201]:
                raise Exception(f"创建会话失败：{response.status_code} - {response.text}")

            data = response.json()
            elapsed = time.time() - start_time

            # 记录请求详情
            self._add_request_detail(result, "create_session", "POST", "/v1/rag/sessions", body, data, elapsed, response.status_code)

            return SessionInfo(
                session_id=data["session_id"],
                title=data["title"],
                pinned=data.get("pinned", False),
                created_at=data.get("created_at"),
                updated_at=data.get("updated_at"),
            )

        finally:
            if client:
                await client.close()

    async def get_session(self, session_id: str, result: QAResult = None) -> Dict[str, Any]:
        """
        获取会话信息

        Args:
            session_id: 会话 ID
            result: 用于记录请求详情的结果对象

        Returns:
            包含 session 和 messages 的字典
        """
        client = None
        start_time = time.time()
        try:
            auth_header = await self.auth_manager.get_auth_header()
            endpoint = self._get_endpoint("rag_get_session")
            if endpoint is None:
                raise Exception("未找到端点配置：rag_get_session，请检查 config/endpoints.yaml")

            client = HttpClient(base_url=self.config.get_base_url(endpoint.base))
            path = self._format_path(endpoint.path, session_id=session_id)

            response = await client.request(
                method=endpoint.method,
                path=path,
                headers=auth_header,
            )

            if response.status_code != 200:
                raise Exception(f"获取会话失败：{response.status_code} - {response.text}")

            data = response.json()
            elapsed = time.time() - start_time

            # 记录请求详情
            self._add_request_detail(result, "get_session", "GET", path, {}, data, elapsed, response.status_code)

            return data

        finally:
            if client:
                await client.close()

    async def update_session_title(self, session_id: str, title: str, result: QAResult = None) -> SessionInfo:
        """
        更新会话标题

        Args:
            session_id: 会话 ID
            title: 新标题（会自动截取到 50 字符，避免数据库报错）
            result: 用于记录请求详情的结果对象

        Returns:
            SessionInfo 更新后的会话信息
        """
        # 截取 title 到 50 字符（避免数据库字段过长）
        truncated_title = title[:50] if len(title) > 50 else title

        start_time = time.time()
        client = None
        try:
            auth_header = await self.auth_manager.get_auth_header()
            endpoint = self._get_endpoint("rag_update_session")
            if endpoint is None:
                raise Exception("未找到端点配置：rag_update_session，请检查 config/endpoints.yaml")

            client = HttpClient(base_url=self.config.get_base_url(endpoint.base))
            path = self._format_path(endpoint.path, session_id=session_id)
            # 只设置 title 字段，不设置 pinned
            body = {"title": truncated_title}

            response = await client.request(
                method=endpoint.method,
                path=path,
                headers=auth_header,
                json=body,
            )

            if response.status_code != 200:
                logger.error(f"响应内容：{response.text}")
                raise Exception(f"更新会话失败：{response.status_code} - {response.text}")

            data = response.json()
            elapsed = time.time() - start_time

            # 记录请求详情
            self._add_request_detail(result, "update_session_title", "PATCH", path, body, data, elapsed, response.status_code)

            session_info = SessionInfo(
                session_id=data["session_id"],
                title=data["title"],
                pinned=data.get("pinned", False),
                created_at=data.get("created_at"),
                updated_at=data.get("updated_at"),
            )
            return session_info

        finally:
            if client:
                await client.close()

    async def create_message(
        self,
        session_id: str,
        content: str,
        knowledge_base_id: str = "ALL_KB",
        scope_mode: str = "all",
        result: QAResult = None,
    ) -> MessageInfo:
        """
        创建提问消息

        Args:
            session_id: 会话 ID
            content: 问题内容
            knowledge_base_id: 知识库 ID
            scope_mode: 范围模式
            result: 用于记录请求详情的结果对象

        Returns:
            MessageInfo 消息信息
        """
        start_time = time.time()
        client = None
        try:
            auth_header = await self.auth_manager.get_auth_header()
            endpoint = self._get_endpoint("rag_create_message")
            if endpoint is None:
                raise Exception("未找到端点配置：rag_create_message，请检查 config/endpoints.yaml")

            client = HttpClient(base_url=self.config.get_base_url(endpoint.base))
            path = self._format_path(endpoint.path, session_id=session_id)
            body = self._build_body(
                endpoint,
                content=content,
                knowledge_base_id=knowledge_base_id,
                scope_mode=scope_mode,
            )

            response = await client.request(
                method=endpoint.method,
                path=path,
                headers=auth_header,
                json=body,
            )

            if response.status_code != 201:
                logger.error(f"响应内容：{response.text}")
                raise Exception(f"创建消息失败：{response.status_code} - {response.text}")

            data = response.json()
            elapsed = time.time() - start_time

            # 记录请求详情
            self._add_request_detail(result, "create_message", "POST", path, body, data, elapsed, response.status_code)

            message_info = MessageInfo(
                message_id=data["message_id"],
                session_id=data["session_id"],
                role=data["role"],
                content=data["content"],
                created_at=data.get("created_at"),
            )
            return message_info

        finally:
            if client:
                await client.close()

    async def create_assistant_message(
        self,
        session_id: str,
        answer: str,
        citations: List[dict],
        knowledge_base_id: str = "ALL_KB",
        scope_mode: str = "all",
        result: QAResult = None,
    ) -> MessageInfo:
        """
        步骤 6：创建助手回复消息（更新提问）

        Args:
            session_id: 会话 ID
            answer: 步骤 5 生成的答案
            citations: 步骤 5 流式响应中 event=done 返回的 citations
            knowledge_base_id: 知识库 ID
            scope_mode: 范围模式
            result: 用于记录请求详情的结果对象

        Returns:
            MessageInfo 消息信息
        """
        start_time = time.time()
        client = None
        try:
            auth_header = await self.auth_manager.get_auth_header()
            # 使用与步骤 4 相同的端点配置
            endpoint = self._get_endpoint("rag_create_message")
            if endpoint is None:
                raise Exception("未找到端点配置：rag_create_message，请检查 config/endpoints.yaml")

            client = HttpClient(base_url=self.config.get_base_url(endpoint.base))
            path = self._format_path(endpoint.path, session_id=session_id)

            # 构建请求体：role="assistant"，其他参数与步骤 4 相同
            body = {
                "role": "assistant",
                "content": answer,
                "citations": citations,
                "scope_mode": scope_mode,
                "knowledge_base_id": knowledge_base_id,
                "doc_ids": [],
                "chat_mode": "pipeline",
            }

            response = await client.request(
                method=endpoint.method,
                path=path,
                headers=auth_header,
                json=body,
            )

            if response.status_code != 201:
                logger.error(f"响应内容：{response.text}")
                raise Exception(f"创建助手消息失败：{response.status_code} - {response.text}")

            data = response.json()
            elapsed = time.time() - start_time

            # 记录请求详情
            self._add_request_detail(result, "create_assistant_message", "POST", path, body, data, elapsed, response.status_code)

            message_info = MessageInfo(
                message_id=data["message_id"],
                session_id=data["session_id"],
                role=data["role"],
                content=data["content"],
                created_at=data.get("created_at"),
            )
            return message_info

        finally:
            if client:
                await client.close()

    def _parse_streaming_response(self, lines: List[str], alert_keywords: List[str] = None) -> FinalAnswer:
        """
        解析流式响应，按优先级提取最终答复（与前端 UI 对齐）
        路径优先级：
          A. intent_email 事件 → 答复 = event.body，type = intent_email
          B. done.status == "no_answer" 且无 token → 答复 = 信封模板，type = no_answer
          C. 聚合 token 后匹配 no-answer 关键词 → 答复 = 信封模板，type = no_answer
          D. 正常聚合 token → 答复 = 聚合文本，type = rag_answer

        Args:
            lines: 流式响应的行列表
            alert_keywords: 异常检测关键字列表

        Returns:
            FinalAnswer 结构
        """
        answer_parts = []
        token_count = 0
        current_event = None
        done_status = None
        intent_email_data = None

        for line in lines:
            stripped = line.strip()
            if not stripped:
                continue

            # 解析 SSE 格式：event: xxx 或 data: xxx
            if stripped.startswith("event: "):
                current_event = stripped[7:].strip()
            elif stripped.startswith("data: "):
                json_str = stripped[6:].strip()
                try:
                    data = json.loads(json_str)

                    # 路径 A：intent_email 事件
                    if current_event == "intent_email" and "body" in data:
                        intent_email_data = data
                        continue  # 继续处理 done 事件

                    # 记录 done 状态
                    if current_event == "done":
                        done_status = data.get("status")

                    # 拼接 token 事件的 delta
                    elif current_event == "token" and "delta" in data:
                        delta = data["delta"]
                        answer_parts.append(delta)
                        token_count += 1
                        # 检测异常关键字
                        if alert_keywords:
                            for keyword in alert_keywords:
                                if keyword in delta:
                                    logger.warning(f"检测到异常关键字 '{keyword}'：{delta[:200]}...")
                                    break

                except (json.JSONDecodeError, ValueError) as e:
                    logger.debug(f"解析 data 失败：{json_str[:100]}... 错误：{e}")

        raw_answer = "".join(answer_parts)

        # 路径 A：intent_email 事件
        if intent_email_data:
            body = intent_email_data.get("body", "")
            lang = intent_email_data.get("lang", "zh-hk")
            intent_quote = intent_email_data.get("intent_quote", {})
            # 从 intent_quote.intent_name 提取引用文档（一定记录）
            intent_citations = self._parse_intent_name(intent_quote)
            # 如果 body 是 no-answer 信封，记录告警但仍记录引用文档
            if body.strip() == ENVELOPE_TEMPLATE_ZH.strip():
                logger.warning(
                    f"[{question_id}] intent_email 事件，但 body 为 no-answer 信封，"
                    f"仍记录 {len(intent_citations)} 条引用文档"
                )
            return FinalAnswer(
                answer_type="intent_email",
                display_body=body,
                plain_text=self._strip_envelope(body, lang),
                lang=lang,
                intent_quote=intent_quote if isinstance(intent_quote, dict) else {},
                citations=intent_citations,
                raw_answer=raw_answer
            )

        # 路径 B：done.status == "no_answer" 且无 token
        if done_status == "no_answer" and token_count == 0:
            return FinalAnswer(
                answer_type="no_answer",
                display_body=ENVELOPE_TEMPLATE_ZH,
                plain_text="抱歉，未能找到相關資料",
                lang="zh-hk",
                intent_quote={},
                citations=[],
                raw_answer=raw_answer
            )

        # 路径 C：聚合 token 后匹配 no-answer 关键词
        if raw_answer and self._contains_no_answer(raw_answer):
            return FinalAnswer(
                answer_type="no_answer",
                display_body=ENVELOPE_TEMPLATE_ZH,
                plain_text="抱歉，未能找到相關資料",
                lang="zh-hk",
                intent_quote={},
                citations=[],
                raw_answer=raw_answer
            )

        # 路径 D：正常 RAG 回答
        return FinalAnswer(
            answer_type="rag_answer",
            display_body=raw_answer,
            plain_text=raw_answer,
            lang="zh-hk",
            intent_quote={},
            citations=[],
            raw_answer=raw_answer
        )

    def _parse_intent_name(self, intent_quote: Dict[str, Any]) -> List[Dict[str, str]]:
        """
        从 intent_quote.intent_name 提取引用文档列表

        Args:
            intent_quote: {"intent_id": "", "intent_name": "名称1, 名称2, ..."}

        Returns:
            [{"name": "名称1"}, {"name": "名称2"}, ...]
        """
        citations = []
        intent_name = intent_quote.get("intent_name", "") if isinstance(intent_quote, dict) else ""
        if intent_name:
            # 去掉 "Possible Intent: " 前缀
            if intent_name.startswith("Possible Intent: "):
                intent_name = intent_name[17:]
            # 按逗号分割，每项作为一条引用
            intent_items = [item.strip() for item in intent_name.split(",") if item.strip()]
            for item in intent_items:
                citations.append({"name": item})
        return citations

    def _contains_no_answer(self, text: str) -> bool:
        """检查文本是否包含 no-answer 关键词"""
        for pattern in NO_ANSWER_PATTERNS:
            if pattern in text:
                return True
        return False

    def _strip_envelope(self, text: str, lang: str = "zh-hk") -> str:
        """剥掉信封头尾，返回纯文字"""
        if lang == "en-us":
            # 英文信封格式
            prefix = "Dear Sir/Madam,\n\n"
            suffix = "\n\nHong Kong Customs"
        else:
            # 中文信封格式
            prefix = "敬啟者：\n\n"
            suffix = "\n\n香港海關"

        if text.startswith(prefix) and text.endswith(suffix):
            return text[len(prefix):-len(suffix)].strip()
        return text

    async def streaming_query(
        self,
        query: str,
        session_id: str,
        knowledge_base_id: str = "ALL_KB",
        top_k: int = None,
        rerank: bool = None,
        citation_mode: str = None,
        question_id: str = None,
    ) -> QAResult:
        """
        流式查询并获取答案

        Args:
            query: 问题
            session_id: 会话 ID
            knowledge_base_id: 知识库 ID
            top_k: 返回的顶部结果数（None 则从配置读取）
            rerank: 是否重排序（None 则从配置读取）
            citation_mode: 引用模式（None 则从配置读取）
            question_id: 问题编号（用于日志追踪）

        Returns:
            QAResult 问答结果
        """
        result = QAResult(question=query, session_id=session_id, question_id=question_id)
        client = None

        try:
            auth_header = await self.auth_manager.get_auth_header()
            endpoint = self._get_endpoint("rag_query_stream")
            if endpoint is None:
                raise Exception("未找到端点配置：rag_query_stream，请检查 config/endpoints.yaml")

            # 根据 rag_create_message 的 chat_mode 决定流式查询接口路径
            chat_mode = self._get_chat_mode()
            if chat_mode == "pipeline":
                query_path = "/v1/rag/query/stream"
            else:
                query_path = "/v1/rag/flex/query/stream"

            # 从 rag_query_stream 配置中读取默认参数
            cfg_body = endpoint.body or {}

            # 优先使用运行时参数，其次使用配置值，最后使用代码默认值
            model = cfg_body.get("model", "qwen3.5-27b-local")
            enable_thinking = cfg_body.get("enable_thinking", False)
            effective_top_k = top_k if top_k is not None else cfg_body.get("top_k", 8)
            effective_rerank = rerank if rerank is not None else cfg_body.get("rerank", True)
            effective_citation_mode = citation_mode if citation_mode is not None else cfg_body.get("citation_mode", "inline")
            doc_ids = cfg_body.get("doc_ids", [])
            scope_mode = cfg_body.get("scope_mode", "all")
            alert_keywords = cfg_body.get("alert_keywords", [])

            # 使用更长的超时时间（流式查询可能需要 2-5 分钟）
            client = HttpClient(
                base_url=self.config.get_base_url(endpoint.base),
                timeout=1200.0,  # 20 分钟超时
            )

            # 根据 chat_mode 构建不同的请求体（参数都从配置读取）
            if chat_mode == "flex":
                body = {
                    "query": query,
                    "session_id": session_id,
                    "model": model,
                    "enable_thinking": enable_thinking,
                    "knowledge_base_id": knowledge_base_id,
                    "doc_ids": doc_ids,
                    "top_k": effective_top_k,
                    "rerank": effective_rerank,
                    "citation_mode": effective_citation_mode,
                    "scope_mode": scope_mode,
                }
            else:  # pipeline
                body = {
                    "query": query,
                    "top_k": effective_top_k,
                    "rerank": effective_rerank,
                    "stream": cfg_body.get("stream", True),
                    "citation_mode": effective_citation_mode,
                    "knowledge_base_id": knowledge_base_id,
                    "session_id": session_id,
                    "model": model,
                    "enable_thinking": enable_thinking,
                }

            start_time = time.time()

            base_url = self.config.get_base_url(endpoint.base)
            full_url = f"{base_url.rstrip('/')}{query_path}"

            # 记录请求到文件
            request_logger.info(f"=== REQUEST ===")
            request_logger.info(f"问题ID: {result.question_id}")
            request_logger.info(f"问题: {query}")
            request_logger.info(f"URL: POST {full_url}")
            request_logger.info(f"Headers: {json.dumps(dict(auth_header), ensure_ascii=False)}")
            request_logger.info(f"Body: {json.dumps(body, ensure_ascii=False, indent=2)}")

            response = await client.request(
                method=endpoint.method,
                path=query_path,
                headers=auth_header,
                json=body,
            )

            if response.status_code != 200:
                logger.error(f"响应内容：{response.text}")
                result.success = False
                result.error = f"请求失败：{response.status_code} - {response.text}"
                elapsed = time.time() - start_time
                self._add_request_detail(result, "streaming_query", "POST", query_path, body, {"error": response.text}, elapsed, response.status_code)
                return result

            # 收集流式响应行
            lines = []
            async for line in response.aiter_lines():
                if line.strip():
                    lines.append(line)

            result.response_time = time.time() - start_time

            # 解析响应
            final_answer = self._parse_streaming_response(lines, alert_keywords)
            result.answer = final_answer.display_body
            result.answer_type = final_answer.answer_type

            # 提取 citations（从 event: done 的 data 中解析）
            citations = []
            metadata = {}
            done_event_data = None
            in_done_event = False
            stream_error = None  # 从任何 event 中捕获的错误

            for i, line in enumerate(lines):
                stripped = line.strip()
                if not stripped:
                    continue

                # 检测 error 事件：event: error 后面紧跟 data: {...}
                if stripped.startswith("event: error"):
                    if i + 1 < len(lines):
                        next_stripped = lines[i + 1].strip()
                        if next_stripped.startswith("data: "):
                            try:
                                err_data = json.loads(next_stripped[6:].strip())
                                stream_error = err_data.get("error", err_data)
                                if isinstance(stream_error, dict):
                                    stream_error = {
                                        "code": stream_error.get("code", ""),
                                        "message": stream_error.get("message", str(stream_error)),
                                    }
                                continue
                            except (json.JSONDecodeError, ValueError):
                                stream_error = {"code": "", "message": next_stripped[6:].strip()}
                                continue

                if stripped.startswith("event: done"):
                    in_done_event = True
                elif in_done_event and stripped.startswith("data: "):
                    try:
                        json_str = stripped[6:].strip()
                        data = json.loads(json_str)
                        citations = data.get("citations", [])
                        metadata = data
                        done_event_data = data
                    except (json.JSONDecodeError, ValueError) as e:
                        logger.debug(f"解析 done 事件失败：{e}")
                    break
                elif stripped.startswith("event:"):
                    in_done_event = False

            # 检测 done 事件中的错误
            if done_event_data and not stream_error:
                # 检测 partial 状态（LLM 未生成答案，后端用 HTTP 200 包装了 LLM 失败）
                done_status = done_event_data.get("status", "")
                if done_status == "partial":
                    # 从 metadata 中提取原始错误信息（如果有 event: error 的话）
                    # partial 状态下，原始错误可能在 metadata 或之前的 event 中
                    # 如果没有具体错误，保留 done 事件全部内容供排查
                    raw_error = json.dumps(done_event_data, ensure_ascii=False)
                    error_msg = f"LLM 未生成回答（后端返回 partial）\n{raw_error}"
                    logger.error(f"流式查询返回 partial 状态：{raw_error}")
                    stream_error = {"code": "PARTIAL", "message": error_msg}
                    result.answer_type = "error"  # 标记为错误类型
                else:
                    error_info = done_event_data.get("error", None)
                    error_code = done_event_data.get("error_code", None) or done_event_data.get("code", None)
                    if error_info:
                        error_msg = error_info if isinstance(error_info, str) else error_info.get("message", str(error_info))
                        stream_error = {"code": error_code, "message": error_msg}
                    elif error_code:
                        error_msg = done_event_data.get("error_message", "") or done_event_data.get("message", "")
                        stream_error = {"code": error_code, "message": error_msg}

            # 检测答案内容是否是错误 JSON（兜底）
            if not stream_error and final_answer.display_body:
                try:
                    answer_json = json.loads(final_answer.display_body)
                    if "error" in answer_json:
                        err = answer_json["error"]
                        error_msg = err.get("message", str(err)) if isinstance(err, dict) else str(err)
                        error_code = err.get("code", "") if isinstance(err, dict) else ""
                        stream_error = {"code": error_code, "message": error_msg}
                    # 检测直接嵌套的错误结构（如 DashScope 格式）
                    elif isinstance(answer_json, dict) and "error" in answer_json and isinstance(answer_json["error"], dict):
                        err = answer_json["error"]
                        stream_error = {"code": err.get("code", ""), "message": err.get("message", str(err))}
                except (json.JSONDecodeError, ValueError):
                    pass

                # 检测答案内容中是否包含已知错误关键词（如错误被嵌入 delta 中）
                if not stream_error:
                    error_keywords = ["quota exceeded", "throttling", "rate limit", "too many requests",
                                      "DashScope returned HTTP", "hour allocated quota exceeded"]
                    for keyword in error_keywords:
                        if keyword.lower() in final_answer.display_body.lower():
                            stream_error = {"code": "", "message": final_answer.display_body.strip()[:500]}
                            break

            # 如果有错误，标记失败并返回
            if stream_error:
                error_code = stream_error.get("code", "")
                error_msg = stream_error.get("message", str(stream_error))
                logger.error(f"流式查询返回错误：{error_code} - {error_msg}")
                result.success = False
                result.error = f"{error_code}: {error_msg}" if error_code else error_msg
                result.answer_type = "error"
                # 将原始错误信息存入 metadata 供 Excel 写入时提取
                result.metadata = {
                    "stream_error": {
                        "code": error_code,
                        "message": error_msg,
                        "raw_error": done_event_data if done_event_data else stream_error,
                    }
                }
                # 错误时使用 FinalAnswer 中的 citations
                if final_answer.citations:
                    result.citations = final_answer.citations
                else:
                    result.citations = citations
                elapsed = time.time() - start_time
                raw_response = "\n".join(lines)
                self._add_request_detail(result, "streaming_query", "POST", query_path, body, {"raw": raw_response, "answer": final_answer.display_body, "citations": citations, "metadata": metadata}, elapsed, response.status_code)

                # 记录错误响应到文件
                request_logger.info(f"=== RESPONSE (ERROR) ===")
                request_logger.info(f"问题ID: {result.question_id}")
                request_logger.info(f"错误: {error_code} - {error_msg}")
                request_logger.info(f"响应状态码: {response.status_code}")
                request_logger.info(f"响应内容: {raw_response}")
                request_logger.info("=" * 50)

                return result

            # 使用 FinalAnswer 中的 citations
            # no_answer 类型不需要提取引用文档
            if final_answer.answer_type == "no_answer":
                result.citations = []
            elif final_answer.citations:
                result.citations = final_answer.citations
            elif citations:
                result.citations = citations
            else:
                result.citations = []
            result.metadata = metadata

            # 记录请求详情 - 保存原始流式响应
            elapsed = time.time() - start_time
            raw_response = "\n".join(lines)  # 原始 SSE 格式响应
            self._add_request_detail(result, "streaming_query", "POST", query_path, body, {"raw": raw_response, "answer": final_answer.display_body, "citations": citations, "metadata": metadata}, elapsed, response.status_code)

            # 记录成功响应到文件
            request_logger.info(f"=== RESPONSE (SUCCESS) ===")
            request_logger.info(f"问题ID: {result.question_id}")
            request_logger.info(f"答复类型: {result.answer_type}")
            request_logger.info(f"答复长度: {len(result.answer) if result.answer else 0}")
            request_logger.info(f"答复内容: {(result.answer or '')[:200]!r}")
            request_logger.info(f"引用文档数: {len(result.citations)}")

            # 始终记录异常情况的详细信息
            if result.citations and result.answer:
                # 检查是否包含 no-answer 相关内容
                no_answer_keywords = ["抱歉", "未能找到", "相關資料"]
                has_no_answer = any(kw in result.answer for kw in no_answer_keywords)
                if has_no_answer:
                    request_logger.warning(f"【异常】答复包含 no-answer 内容但仍有引用文档，请检查")
                    request_logger.warning(f"完整答复: {result.answer!r}")
                    request_logger.warning(f"引用文档: {result.citations}")
                    # 如果开启了 debug_sse_log，额外记录完整 SSE 响应
                    if self._get_debug_sse_log():
                        request_logger.warning(f"完整 SSE 响应: {raw_response}")

            # 如果开启了 debug_sse_log，额外记录完整 SSE 响应
            if self._get_debug_sse_log():
                request_logger.info(f"【DEBUG】完整 SSE 响应:")
                request_logger.info(raw_response)

            request_logger.info("=" * 50)

            result.success = True
            return result

        except Exception as e:
            logger.error(f"流式查询失败：{e}")
            result.success = False
            result.error = str(e)
            return result

        finally:
            if client:
                await client.close()

    async def run_single_qa_test(
        self,
        question: str,
        knowledge_base_id: str = "ALL_KB",
        session_title: str = None,
        question_id: str = None,
    ) -> QAResult:
        """
        执行单次完整的问答测试流程

        Args:
            question: 问题
            knowledge_base_id: 知识库 ID
            session_title: 会话标题（从测试集读取）
            question_id: 问题编号（从测试集第一列读取）

        Returns:
            QAResult 问答结果
        """
        result = QAResult(question=question, question_id=question_id)

        try:
            # 1. 创建新会话
            session_info = await self.create_session("New Chat", result)
            session_id = session_info.session_id
            result.session_id = session_id

            # 2. 获取会话详情（验证会话创建成功）
            await self.get_session(session_id, result)

            # 3. 更新会话标题（如果提供了标题）
            if session_title:
                await self.update_session_title(session_id, session_title, result)

            # 4. 创建提问消息
            message_info = await self.create_message(session_id, question, knowledge_base_id, result=result)
            result.message_id = message_info.message_id

            # 5. 流式查询获取答案
            qa_result = await self.streaming_query(question, session_id, knowledge_base_id, question_id=question_id)
            result.answer = qa_result.answer
            result.answer_type = qa_result.answer_type
            result.response_time = qa_result.response_time
            result.metadata = qa_result.metadata
            result.citations = qa_result.citations
            result.success = qa_result.success
            result.error = qa_result.error
            # 合并 streaming_query 的请求详情
            result.request_details.extend(qa_result.request_details)

            # 6. 创建助手回复消息（更新提问）
            if qa_result.answer and qa_result.citations is not None:
                await self.create_assistant_message(
                    session_id=session_id,
                    answer=qa_result.answer,
                    citations=qa_result.citations,
                    knowledge_base_id=knowledge_base_id,
                    result=result,
                )

            return result

        except Exception as e:
            result.success = False
            result.error = str(e)
            return result

    async def run_batch_qa_tests(
        self,
        questions: List[tuple],  # [(question, title, question_id), ...]
        knowledge_base_id: str = "ALL_KB",
        max_concurrent: int = 3,
    ) -> BatchQAResult:
        """
        批量执行问答测试

        Args:
            questions: 问题列表，每个元素为 (问题，标题，编号) 元组
            knowledge_base_id: 知识库 ID
            max_concurrent: 最大并发数

        Returns:
            BatchQAResult 批量问答结果
        """
        result = BatchQAResult(start_time=datetime.now())
        result.total = len(questions)

        if result.total == 0:
            result.end_time = datetime.now()
            return result

        logger.info("=" * 60)
        logger.info(f"开始执行问答测试，共 {result.total} 个问题，最大并发数：{max_concurrent}")
        logger.info("=" * 60)

        # 并发控制
        semaphore = asyncio.Semaphore(max_concurrent)
        interrupted = False  # 标记是否被中断

        async def run_test_with_semaphore(idx: int, q: tuple) -> QAResult:
            async with semaphore:
                if len(q) >= 3:
                    question, title, question_id = q
                else:
                    question, title = q
                    question_id = None
                logger.info(f"[{idx + 1}/{result.total}] 开始处理：{question[:50]}...")
                qa_result = await self.run_single_qa_test(
                    question=question,
                    knowledge_base_id=knowledge_base_id,
                    session_title=title,
                    question_id=question_id,
                )
                status = "成功" if qa_result.success else "失败"
                logger.info(f"[{idx + 1}/{result.total}] 完成：{status}, 耗时：{qa_result.response_time:.2f}s")
                return qa_result

        # 创建所有任务
        tasks = [asyncio.ensure_future(run_test_with_semaphore(idx, q)) for idx, q in enumerate(questions)]

        try:
            # 使用 asyncio.wait 等待所有任务完成，支持中断时保留已完成的结果
            pending = set(tasks)
            completed_results: List[QAResult] = []

            while pending:
                done, pending = await asyncio.wait(pending, return_when=asyncio.FIRST_COMPLETED)
                for task in done:
                    if not task.cancelled() and not task.exception():
                        completed_results.append(task.result())
                    elif task.exception():
                        logger.error(f"任务异常：{task.exception()}")

            # 正常完成
            result.results = completed_results

        except (asyncio.CancelledError, KeyboardInterrupt):
            # 中断时收集已完成任务的结果
            logger.warning("测试被中断，保存已完成的测试结果...")
            interrupted = True
            for task in tasks:
                if task.done() and not task.cancelled() and not task.exception():
                    try:
                        r = task.result()
                        if r not in completed_results:
                            completed_results.append(r)
                    except Exception:
                        pass
                else:
                    task.cancel()
            result.results = completed_results

        # 统计结果（包括中断时的部分结果）
        result.success = sum(1 for r in result.results if r.success)
        result.failed = sum(1 for r in result.results if not r.success)
        result.end_time = datetime.now()

        logger.info("=" * 60)
        if interrupted:
            logger.info(f"测试被中断：已完成 {len(result.results)}/{result.total}, 成功 {result.success}, 失败 {result.failed}")
        else:
            logger.info(f"测试完成：成功 {result.success}/{result.total}, 成功率：{result.success_rate:.1%}")
        logger.info(f"平均响应时间：{result.avg_response_time:.2f}s, P95: {result.p95_response_time:.2f}s")
        logger.info("=" * 60)

        return result

    async def load_questions_from_excel(
        self,
        testset_path: str,
        question_column: int = 2,
        title_column: int = None,  # 可选，用于读取标题
        id_column: int = 1,  # 编号列（1-based）
        sheet_name: str = None,  # Excel 工作表名称（None 表示使用默认第一个工作表）
        start_row: int = 2,
        end_row: int = None,
    ) -> List[tuple]:
        """
        从 Excel 文件加载问题列表

        Args:
            testset_path: Excel 文件路径
            question_column: 问题所在列（1-based）
            title_column: 标题所在列（1-based），None 则使用问题作为标题
            id_column: 编号所在列（1-based），默认第 1 列
            sheet_name: Excel 工作表名称（None 表示使用默认第一个工作表）
            start_row: 起始行（1-based），跳过表头
            end_row: 结束行（None 表示到最后）

        Returns:
            问题列表 [(question, title, question_id), ...]
        """
        try:
            import openpyxl
        except ImportError:
            raise ImportError("需要安装 openpyxl: pip install openpyxl")

        wb = openpyxl.load_workbook(testset_path, read_only=True)

        # 根据 sheet_name 获取工作表
        if sheet_name:
            ws = wb[sheet_name]
        else:
            ws = wb.active

        questions = []
        end_row_param = end_row if end_row else ws.max_row

        for row in ws.iter_rows(min_row=start_row, max_row=end_row_param, min_col=1, max_col=max(question_column, title_column or 1, id_column)):
            question_cell = row[question_column - 1] if question_column <= len(row) else None
            title_cell = row[title_column - 1] if title_column and title_column <= len(row) else None
            id_cell = row[id_column - 1] if id_column <= len(row) else None

            question = str(question_cell.value).strip() if question_cell and question_cell.value else None
            question_id = str(id_cell.value).strip() if id_cell and id_cell.value else None

            # 如果没有指定 title_column，使用问题作为标题
            if title_column and title_cell and title_cell.value:
                title = str(title_cell.value).strip()
            else:
                title = question

            if question:
                questions.append((question, title, question_id))

        wb.close()
        return questions

    async def save_results_to_excel(
        self,
        results: BatchQAResult,
        output_path: str,
        template_path: str = None,
        resolve_doc_names: bool = True,  # 是否解析文档名称
    ) -> str:
        """
        将问答结果保存到 Excel

        Args:
            results: 批量问答结果
            output_path: 输出文件路径
            template_path: 模板文件路径（可选，用于保留原始问题数据）

        Returns:
            输出文件路径
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
        except ImportError:
            raise ImportError("需要安装 openpyxl: pip install openpyxl")

        # 如果提供了模板文件，加载它
        if template_path and Path(template_path).exists():
            wb = openpyxl.load_workbook(template_path)
        else:
            wb = openpyxl.Workbook()

        ws = wb.active
        ws.title = "问答测试结果"

        # 定义样式
        header_font = Font(name='Microsoft YaHei Light', size=8)
        cell_font = Font(name='Microsoft YaHei Light', size=8)
        cell_alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

        # 设置表头（第 1 行）：编号 | 提问 | 生成回答 | 响应时间 | 引用文档
        # 根据 rag_create_message 的 chat_mode 配置动态设置第 3 列表头
        chat_mode_header = self._get_chat_mode_header()
        headers = ["编号", "提问", chat_mode_header, "响应时间", "引用文档"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.alignment = cell_alignment

        # 填充数据（从第 2 行开始）
        # 如果需要解析文档名称，先收集所有文档 ID
        doc_name_cache = {}
        if resolve_doc_names:
            all_doc_ids = set()
            for result in results.results:
                citations = result.citations or (result.metadata.get('citations', []) if result.metadata else [])
                for cite in citations:
                    if isinstance(cite, dict):
                        doc_id = cite.get('doc_id', '')
                        if doc_id:
                            all_doc_ids.add(doc_id)

            # 批量查询文档名称
            if all_doc_ids:
                logger.info(f"正在查询 {len(all_doc_ids)} 个文档名称...")
                for doc_id in all_doc_ids:
                    detail = await self.get_document_detail(doc_id)
                    if detail:
                        doc_name = detail.get('doc_name', '')
                        if doc_name:
                            # 用原始 doc_id 作为 key，确保查找一致
                            doc_name_cache[doc_id] = doc_name
                        else:
                            logger.debug(f"文档 {doc_id} 查询成功但无 doc_name: {detail}")
                    else:
                        logger.warning(f"文档 {doc_id} 查询失败（可能不存在或无权限）")
                logger.info(f"已缓存 {len(doc_name_cache)}/{len(all_doc_ids)} 个文档名称")

        # 按原始编号排序（按 question_id 的自然顺序）
        sorted_results = sorted(results.results, key=lambda r: r.question_id or "")

        for row_idx, result in enumerate(sorted_results, 2):
            # 第 1 列：编号
            id_cell = ws.cell(row=row_idx, column=1, value=result.question_id or "")
            id_cell.font = cell_font
            id_cell.alignment = cell_alignment

            # 第 2 列：提问
            question_cell = ws.cell(row=row_idx, column=2, value=result.question)
            question_cell.font = cell_font
            question_cell.alignment = cell_alignment

            # 第 3 列：生成回答
            answer_cell = ws.cell(row=row_idx, column=3, value=result.answer or "")
            answer_cell.font = cell_font
            answer_cell.alignment = cell_alignment

            # 第 4 列：响应时间
            time_cell = ws.cell(row=row_idx, column=4, value=f"{result.response_time:.2f} 秒" if result.response_time else "")
            time_cell.font = cell_font
            time_cell.alignment = cell_alignment

            # 第 5 列：引用文档（从 citations 中提取）
            # no_answer 类型不需要提取引用文档
            citations_text = ""
            answer_type = getattr(result, 'answer_type', '') or ''
            if answer_type != "no_answer" and result.citations:
                citation_lines = []
                for cite in result.citations:
                    if isinstance(cite, dict):
                        # 优先使用 name 字段（intent_email 引用）
                        name = cite.get('name', '')
                        if name:
                            ref_id = cite.get('ref_id', '')
                            citation_lines.append((ref_id, f"{ref_id}:{name}" if ref_id else name))
                            continue
                        # 否则使用 doc_id 查找文档名称
                        doc_id = cite.get('doc_id', '')
                        ref_id = cite.get('ref_id', '')
                        if doc_id:
                            doc_name = doc_name_cache.get(doc_id, '')
                            # 只记录文档名称，不记录 doc_id
                            if doc_name:
                                citation_lines.append((ref_id, f"{ref_id}:{doc_name}" if ref_id else doc_name))
                            elif doc_id:
                                citation_lines.append((ref_id, f"{ref_id}:{doc_id}" if ref_id else doc_id))
                    elif isinstance(cite, str):
                        citation_lines.append((None, cite))
                # 按 ref_id 数字排序（没有 ref_id 的放在最后）
                citation_lines.sort(key=lambda x: (0 if x[0] and '_' in x[0] else 1, int(x[0].split('_')[1]) if x[0] and '_' in x[0] else 0))
                citations_text = "\n".join([line[1] for line in citation_lines])

            # 如果 citations_text 为空，且不是 no_answer，尝试从 metadata 中获取
            if not citations_text and answer_type != "no_answer" and result.metadata:
                metadata_citations = result.metadata.get('citations', [])
                if metadata_citations:
                    citation_lines = []
                    for cite in metadata_citations:
                        if isinstance(cite, dict):
                            # 优先使用 name 字段（intent_email 引用）
                            name = cite.get('name', '')
                            if name:
                                ref_id = cite.get('ref_id', '')
                                citation_lines.append((ref_id, f"{ref_id}:{name}" if ref_id else name))
                                continue
                            # 否则使用 doc_id 查找文档名称
                            doc_id = cite.get('doc_id', '')
                            ref_id = cite.get('ref_id', '')
                            if doc_id:
                                doc_name = doc_name_cache.get(doc_id, '')
                                # 只记录文档名称，不记录 doc_id
                                if doc_name:
                                    citation_lines.append((ref_id, f"{ref_id}:{doc_name}" if ref_id else doc_name))
                                elif doc_id:
                                    citation_lines.append((ref_id, f"{ref_id}:{doc_id}" if ref_id else doc_id))
                        elif isinstance(cite, str):
                            citation_lines.append((None, cite))
                    # 按 ref_id 数字排序（没有 ref_id 的放在最后）
                    citation_lines.sort(key=lambda x: (0 if x[0] and '_' in x[0] else 1, int(x[0].split('_')[1]) if x[0] and '_' in x[0] else 0))
                    citations_text = "\n".join([line[1] for line in citation_lines])

            # 测试失败时，在引用文档列记录失败信息
            if not result.success:
                fail_info_parts = []
                # 优先从 metadata 中提取流式查询原始错误
                if result.metadata and "stream_error" in result.metadata:
                    stream_err = result.metadata["stream_error"]
                    raw_err = stream_err.get("raw_error", stream_err)
                    if isinstance(raw_err, dict):
                        fail_info_parts.append(f"【流式查询错误】{json.dumps(raw_err, ensure_ascii=False)}")
                    else:
                        fail_info_parts.append(f"【流式查询错误】{raw_err}")
                elif result.request_details:
                    for req_detail in result.request_details:
                        api_name = req_detail.get("api", "")
                        status = req_detail.get("status_code", "")
                        resp = req_detail.get("response_data", {})
                        elapsed = req_detail.get("elapsed", "")
                        # 记录非成功响应 或 无状态码（网络异常/超时等）
                        if not status or (status != 200 and status != 201):
                            status_label = f"HTTP {status}" if status else "未返回"
                            fail_info_parts.append(f"【{api_name}】{status_label} | 耗时 {elapsed}")
                            if isinstance(resp, dict) and "error" in resp:
                                fail_info_parts.append(f"  响应: {json.dumps(resp, ensure_ascii=False)[:500]}")
                            elif isinstance(resp, dict):
                                fail_info_parts.append(f"  响应: {json.dumps(resp, ensure_ascii=False)[:300]}")
                # 如果没有任何错误记录，记录全局错误
                if not fail_info_parts and result.error:
                    fail_info_parts.append(f"【异常】{result.error}")
                if fail_info_parts:
                    fail_info = "\n".join(fail_info_parts)
                    citations_text = (citations_text + "\n\n" + fail_info) if citations_text else fail_info

            cite_cell = ws.cell(row=row_idx, column=5, value=citations_text)
            cite_cell.font = cell_font
            cite_cell.alignment = cell_alignment

        # 调整列宽
        ws.column_dimensions['A'].width = 10  # 编号
        ws.column_dimensions['B'].width = 40  # 提问
        ws.column_dimensions['C'].width = 80  # 生成回答
        ws.column_dimensions['D'].width = 15  # 响应时间
        ws.column_dimensions['E'].width = 60  # 引用文档

        # 保存文件（处理中文路径）
        output_path = str(output_path)
        wb.save(output_path)
        wb.close()

        return output_path

    async def run_qa_test_from_config(self, scenario_name: str = "hk_customs") -> BatchQAResult:
        """
        从配置加载测试集并执行问答测试

        Args:
            scenario_name: 测试场景名称

        Returns:
            BatchQAResult 批量问答结果
        """
        # 获取场景配置
        scenario = self.config.get_scenario(scenario_name)
        if not scenario or not scenario.qa_test:
            raise ValueError(f"测试场景 '{scenario_name}' 未找到或未配置 qa_test")

        qa_config = scenario.qa_test

        # 加载问题
        questions = await self.load_questions_from_excel(
            testset_path=qa_config.testset_path,
            question_column=qa_config.question_column,
            start_row=qa_config.start_row,
            end_row=qa_config.end_row,
        )

        # 执行测试
        results = await self.run_batch_qa_tests(
            questions=questions,
            knowledge_base_id=qa_config.knowledge_base_id,
            max_concurrent=qa_config.max_concurrent,
        )

        # 保存结果到 Excel
        output_dir = Path(qa_config.testset_path).parent / "results"
        output_dir.mkdir(parents=True, exist_ok=True)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        chat_mode_name = self._get_chat_mode_header()
        output_path = output_dir / f"{chat_mode_name}_{timestamp}.xlsx"

        await self.save_results_to_excel(
            results=results,
            output_path=str(output_path),
            template_path=qa_config.testset_path,
        )

        return results
