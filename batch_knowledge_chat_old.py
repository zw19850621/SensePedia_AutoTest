import re
from datetime import datetime
from openpyxl import load_workbook
import os
import httpx
import asyncio
import json
import sys
from Crypto.Cipher import PKCS1_v1_5
from Crypto.PublicKey import RSA
import base64
import time

# 测试地址
#海关原测试环境
test_url = "https://103.237.28.248:12002"

#海关沐熙环境
#test_url = "https://sensepedia.mooo.com:8800"

# 测试集模板文件（只读）
template_file_name = r"D:\测试项目\Sensepedia\项目版本\海关POC\测试集\测试集全量.xlsx"

# 测试结果输出文件（自动生成，可手动指定）
output_file_name = ""  # 为空时自动生成，格式: 模板名_时间戳.xlsx

# 测试sheet页名称
test_sheet_name = "调试"

# 测试数据起始行
test_case_row = 2

# query所在列（模板中问题所在列）
test_case_column = 2

# 输出文件列配置
output_id_column = 1          # A列: 序号
output_question_column = 2    # B列: 问题（原样复制）
output_answer_column = 3      # C列: 答案
output_hit_column = 4          # D列: 命中信息
output_time_column = 5         # E列: 响应时间


# 每完成多少题增量保存一次（0=禁用增量保存）
INCREMENTAL_SAVE_INTERVAL = 5


# 并发请求数
CONCURRENCY_COUNT = 6

# 登录配置（测试账号的用户名和密码）
USERNAME = "zhangwei" 
PASSWORD = "123456"  

# RSA公钥（用于密码加密,勿改动）
PUBLIC_KEY = "MFwwDQYJKoZIhvcNAQEBBQADSwAwSAJBAKRB3BJD9rbp5UrM1hGQLr+q5PfCzjeXGd+G+5NmPnWFEqXllKA+aQSmJ/+G/IkOthk1qyFIRKUjs+3YcN9dhpUCAwEAAQ=="

# 全局token变量
token = None

# 响应时间统计
response_times = []  # 存储所有请求的响应时间


class RSAEncryptor:
    """RSA加密工具类"""
    
    def __init__(self, public_key_base64):
        """
        初始化RSA加密器
        
        Args:
            public_key_base64: Base64编码的公钥
        """
        self.public_key_base64 = public_key_base64
        self._cipher = None
        self._init_cipher()
    
    def _init_cipher(self):
        """初始化加密器"""
        try:
            # 解码Base64公钥
            public_key_der = base64.b64decode(self.public_key_base64)
            # 导入RSA公钥
            rsa_key = RSA.importKey(public_key_der)
            # 创建加密器
            self._cipher = PKCS1_v1_5.new(rsa_key)
        except Exception as e:
            print(f"初始化RSA加密器失败: {e}")
            raise
    
    def encrypt(self, plaintext):
        """
        RSA加密文本
        
        Args:
            plaintext: 要加密的明文
            
        Returns:
            加密后的十六进制字符串
        """
        try:
            if not self._cipher:
                self._init_cipher()
            
            # 加密数据
            encrypted_data = self._cipher.encrypt(plaintext.encode('utf-8'))
            
            if encrypted_data is None:
                raise ValueError("加密失败，返回None")
            
            # 转换为十六进制字符串
            encrypt_str = encrypted_data.hex()
            return encrypt_str
            
        except Exception as e:
            print(f"加密过程中发生错误: {e}")
            return plaintext  # 加密失败返回原文


async def login_and_get_token(username, password):
    """
    登录并获取token
    
    Args:
        username: 用户名
        password: 明文密码
        
    Returns:
        token字符串或None(登录失败)
    """
    try:
        # 创建RSA加密器
        encryptor = RSAEncryptor(PUBLIC_KEY)
        
        # 加密密码
        encrypted_password = encryptor.encrypt(password)
        
        # 登录请求头
        headers = {
            "Content-Type": "application/json",
            "Accept": "application/json"
        }
        
        # 登录请求体
        json_data = {
            "remeberMe": "true", 
            "username": username,
            "password": encrypted_password
        }
        
        print(f"尝试登录用户: {username}")
        print("正在加密密码并发送登录请求...")
        
        start_time = time.time()
        async with httpx.AsyncClient(timeout=30, verify=False) as client:
            # 发送登录请求
            response = await client.post(
                url=test_url + "/api/v1/login", 
                headers=headers,
                json=json_data
            )
            response_time = time.time() - start_time
            
        # 检查响应状态
        response.raise_for_status()
        
        # 解析响应
        result = response.json()
        
        # 检查登录是否成功
        if result.get("code") == 200 and "token" in result:
            token = result.get("token")
            if token:
                print(f"登录成功！响应时间: {response_time:.2f}秒")
                return token
            else:
                print("登录失败:响应中未找到token")
                print(f"响应内容: {result}")
        else:
            print(f"登录失败：{result.get('msg', '未知错误')}")
            print(f"完整响应: {result}")
            
    except httpx.HTTPError as e:
        print(f"HTTP错误 - 登录失败: {e}")
        if hasattr(e, 'response'):
            print(f"响应状态码: {e.response.status_code}")
            print(f"响应内容: {e.response.text}")
    except json.JSONDecodeError as e:
        print(f"JSON解析错误: {e}")
        print(f"原始响应: {response.text if 'response' in locals() else '无响应'}")
    except Exception as e:
        print(f"登录过程中发生未知错误: {e}")
        import traceback
        traceback.print_exc()
    
    return None


def is_file_locked(filepath):
    """检查文件是否被占用"""
    if not os.path.exists(filepath):
        return False
    
    try:
        fd = os.open(filepath, os.O_RDWR | os.O_EXCL)
        os.close(fd)
        return False
        
    except (OSError, IOError) as e:
        if e.errno in [13, 32]: 
            return True
        return False


def check_excel_file():
    """检查Excel模板文件是否可用"""
    if not os.path.exists(template_file_name):
        print(f"文件'{template_file_name}'不存在！")
        return False

    if is_file_locked(template_file_name):
        print(f"文件'{template_file_name}'被其他程序占用!")
        print("请确保Excel文件已关闭, 然后再运行此脚本!")
        return False

    try:
        wb = load_workbook(template_file_name, read_only=True)
        wb.close()
        return True
    except Exception as e:
        print(f"打开Excel文件失败: {e}")
        return False


def read_excel(sheet):
    """读取Excel中的问题, 过滤空值"""
    questions = []
    row_numbers = []
    question_ids = []  # 从A列读取原始编号

    n = 0
    for row in sheet.iter_rows(min_col=test_case_column, max_col=test_case_column):
        n += 1
        if n < test_case_row:
            continue

        # 同时读取A列的原始编号
        id_cell = sheet.cell(row=n, column=1)
        qid = id_cell.value

        cell = row[0]
        if cell.value is None:
            print(f"第{n}行问题为空，跳过")
            continue

        question = str(cell.value).strip()
        if not question:
            print(f"第{n}行问题为空字符串，跳过")
            continue

        questions.append(question)
        row_numbers.append(n)
        question_ids.append(qid)

    print(f"提取到有效问题数量: {len(questions)}")
    return questions, row_numbers, question_ids


async def create_id(question, headers):
    """创建会话"""
    try:
        payload = {
            "name": question[:50],  
            "scope": 2
        }
        start_time = time.time()
        async with httpx.AsyncClient(timeout=600, verify=False) as client:
            response = await client.post(
                url=test_url + "/api/v1/system/session/add", 
                headers=headers, 
                json=payload
            )
            response_time = time.time() - start_time
            
        response.raise_for_status()
        result = response.json()
        session_id = result.get("data", {}).get("id", "")
        
        if not session_id:
            print(f"创建会话失败, 未获取到session_id")
            return None, response_time
            
        return session_id, response_time
            
    except httpx.HTTPError as e:
        print(f"HTTP错误 - 创建会话失败: {e}")
        return None, 0
    except Exception as e:
        print(f"创建会话失败: {e}")
        return None, 0


async def ask_question(session_id, question, headers):
    """提问并获取答案"""
    payload = {
        "conversationType": 4,
        "knowledgeId": "ALL",
        "preConversationId": None,
        "regenerate": {
            "isRegenerate": False
        },
        "knowledgeScope": {
            "basicAbility": 1,
            "privateDatabase": 1
        },
        "config": {
            "maxTokens": 8192,
            "temperature": 0.1,
            "topP": 0.95,
            "stopSequence": "",
            "repetitionPenalty": 0,
            "frequencyPenalty": 0
        },
        "sessionId": session_id,
        "lang": "default",
        "query": question,
        "onContinue": False,
        "prompt": ""
    }

    try:
        start_time = time.time()
        async with httpx.AsyncClient(timeout=600, verify=False) as client:
            response = await client.post(
                url=test_url + "/api/v1/knowledge/chat",
                headers=headers,
                json=payload
            )
            response_time = time.time() - start_time

        response.raise_for_status()
        result = response.text
        answer, hit_info = extract_final_answer(result)
        return answer, hit_info, response_time

    except httpx.HTTPError as e:
        print(f"HTTP错误 - 提问失败: {e}")
        return f"Error: HTTP请求失败 - {e}", "", 0
    except Exception as e:
        print(f"提问失败: {e}")
        return f"Error: {e}", "", 0


def extract_final_answer(response_text):
    """从响应中提取最终答案和命中信息

    Returns:
        (answer, hit_info) 元组
        answer: 提取的答案文本
        hit_info: 命中信息字符串
            - 知识库问答流程: 命中文档名，逗号分隔
            - 意图识别流程: 命中意图名称，逗号分隔
            - 无命中: 空字符串
    """
    if not response_text:
        print("响应文本为空")
        return "Error: 响应为空", ""

    lines = response_text.strip().split('\n')

    for line in reversed(lines):
        if not line.strip():
            continue

        try:
            data = json.loads(line)

            if "answer" in data and data["answer"]:
                answer = data["answer"].replace("<br>", "\n").replace("<br/>", "\n")
                answer = answer.strip()

                # 解析 items 字段
                items_str = data.get("items", "")
                items = _parse_items(items_str)

                # 判断是否走了知识库问答流程
                ref_ids = re.findall(r'\[ref_(\d+)\]', answer)

                if ref_ids and items:
                    # 知识库问答流程: 按 ref_id 分组展示，格式: ref_1: doc1, doc2\nref_3: doc3
                    lines = []
                    seen_docs = {}  # ref_id_str -> set of doc names
                    for ref_id_str in ref_ids:
                        idx = int(ref_id_str) - 1  # 1-based → 0-based
                        if 0 <= idx < len(items):
                            doc_name = items[idx].get("doc_name", "")
                            # 去掉 "All: " 前缀
                            if doc_name.startswith("All: "):
                                doc_name = doc_name[5:]
                            if doc_name:
                                if ref_id_str not in seen_docs:
                                    seen_docs[ref_id_str] = set()
                                seen_docs[ref_id_str].add(doc_name)
                    for ref_id_str in sorted(seen_docs.keys(), key=int):
                        docs = seen_docs[ref_id_str]
                        lines.append(f"ref_{ref_id_str}: {', '.join(sorted(docs))}")
                    hit_info = "\n".join(lines) if lines else ""
                elif items:
                    # 意图识别流程: 提取 intentName 中 "Possible Intent: " 后的内容
                    hit_info = _extract_intent(items)
                else:
                    hit_info = ""

                return answer, hit_info

        except json.JSONDecodeError:
            continue
        except Exception as e:
            print(f"解析JSON失败: {e}")
            continue

    print("未找到有效的answer字段")
    return "Error: 未找到有效答案", ""


def _parse_items(items_str):
    """解析 items 字段（可能是嵌套JSON字符串）"""
    if not items_str:
        return []
    if isinstance(items_str, list):
        return items_str
    if isinstance(items_str, str):
        try:
            return json.loads(items_str)
        except json.JSONDecodeError:
            return []
    return []


def _extract_intent(items):
    """从意图识别场景的 items 中提取意图名称，按逗号分割后换行展示"""
    intent_parts = []
    for item in items:
        intent_name = item.get("intentName", "")
        if not intent_name:
            continue
        # 去掉 "Possible Intent: " 前缀
        prefix = "Possible Intent: "
        if intent_name.startswith(prefix):
            intent_name = intent_name[len(prefix):]
        if intent_name:
            intent_parts.append(intent_name)
    # 按逗号分割后换行展示
    lines = []
    for part in intent_parts:
        for line in part.split(","):
            line = line.strip()
            if line:
                lines.append(line)
    return "\n".join(lines) if lines else ""


async def process_question(_idx, row_num, total_row, question, semaphore, headers):
    """处理单个问题，包括创建会话和提问

    Returns:
        (question_id, question, answer, hit_info, create_time, ask_time)
    """
    if not question or not question.strip():
        print(f"第{row_num}行: 问题为空 [跳过]")
        return row_num, question, f"问题为空 [跳过]", "", 0, 0

    try:
        async with semaphore:
            # 创建会话
            session_id, create_time = await create_id(question, headers)
            if not session_id:
                return row_num, question, f"Error: 创建会话失败", "", create_time, 0

            # 提问获取答案
            answer, hit_info, ask_time = await ask_question(session_id, question, headers)

            # 计算总响应时间
            total_time = create_time + ask_time

            # 记录响应时间
            response_times.append(total_time)

            if answer and not answer.startswith("Error"):
                print(f"第{row_num}/{total_row}行 [获取答案成功]")
                print(f"{question}\n")
                print(f"{answer}")
                if hit_info:
                    print(f"[命中信息] {hit_info}")

            else:
                print(f"第{row_num}/{total_row}行 [获取答案失败]")
                print(f"{question}\n")
                print(f"{answer}")


            return row_num, question, answer, hit_info, create_time, ask_time

    except asyncio.TimeoutError:
        print(f"第{row_num}行 [超时]")
        return row_num, question, "Error: 请求超时", "", 0, 0
    except Exception as e:
        print(f"第{row_num}行 [异常]")
        return row_num, question, f"Error: {str(e)[:100]}", "", 0, 0


def _write_results_to_sheet(sheet, results, start_row=2):
    """将结果批量写入 sheet

    Args:
        sheet: Excel sheet
        results: 每个元素为 (question_id, question, answer, hit_info, create_time, ask_time)
        start_row: 结果写入起始行（默认第2行）
    """
    row_idx = start_row
    for question_id, question, answer, hit_info, create_time, ask_time in results:
        sheet.cell(row=row_idx, column=output_id_column, value=question_id)
        sheet.cell(row=row_idx, column=output_question_column, value=question)
        sheet.cell(row=row_idx, column=output_answer_column, value=answer)

        total_time = create_time + ask_time
        if total_time > 0:
            sheet.cell(row=row_idx, column=output_time_column, value=f"{total_time:.2f}秒")
        else:
            sheet.cell(row=row_idx, column=output_time_column, value="N/A")

        if hit_info:
            sheet.cell(row=row_idx, column=output_hit_column, value=hit_info)

        row_idx += 1


def _save_results_incremental(wb_out, sheet, completed_results):
    """增量保存已完成的结果到文件"""
    _write_results_to_sheet(sheet, completed_results)
    try:
        wb_out.save(output_file_name)
    except Exception:
        pass  # 增量保存失败不影响主流程


def calculate_statistics():
    """计算响应时间统计"""
    if not response_times:
        return None
    
    valid_times = [t for t in response_times if t > 0]
    
    if not valid_times:
        return None
    
    stats = {
        "total_requests": len(response_times),
        "valid_requests": len(valid_times),
        "avg_time": sum(valid_times) / len(valid_times),
        "min_time": min(valid_times),
        "max_time": max(valid_times),
        "total_time": sum(valid_times)
    }
    
    return stats


async def main():
    """主函数"""
    global token

    # 1. 登录获取token
    print("\n步骤1: 登录获取token")

    if not USERNAME or USERNAME == "your_username":
        print("错误：请先在代码中配置测试用户名&密码")
        input("按回车键退出...")
        sys.exit(1)

    token = await login_and_get_token(USERNAME, PASSWORD)

    if not token:
        print("登录失败，无法继续执行！")
        input("按回车键退出...")
        sys.exit(1)

    # 准备请求头
    headers = {"Authorization": f"Bearer {token}"}
    print("Token已获取")

    # 2. 检查模板文件
    print("\n步骤2: 检查模板文件")

    if not check_excel_file():
        input("按回车键退出...")
        sys.exit(1)

    try:
        # 从模板读取问题（只读）
        wb_template = load_workbook(template_file_name, read_only=True)
        sheet_template = wb_template[test_sheet_name]
        questions, row_numbers, _question_ids = read_excel(sheet_template)
        wb_template.close()

        total_questions = len(questions)

        if not questions:
            print("未找到有效问题，程序退出")
            input("按回车键退出...")
            return

        # 确定输出文件路径
        global output_file_name
        if not output_file_name:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            template_name = os.path.splitext(os.path.basename(template_file_name))[0]
            output_dir = os.path.dirname(template_file_name)
            output_file_name = os.path.join(output_dir, f"{template_name}_{timestamp}.xlsx")

        # 复制模板创建输出文件
        import shutil
        shutil.copy2(template_file_name, output_file_name)
        print(f"输出文件: {output_file_name}")

        # 加载输出文件（可写）
        wb_out = load_workbook(output_file_name)
        sheet_out = wb_out[test_sheet_name]

# 清除结果区域（只清答案/时间/命中列，保留问题列）
        for row in sheet_out.iter_rows(min_row=test_case_row,
                                        min_col=output_answer_column,
                                        max_col=output_time_column):
            for cell in row:
                cell.value = None

        print(f"\n步骤3: 开始处理问题(并发数: {CONCURRENCY_COUNT})")

        # 创建信号量控制并发
        semaphore = asyncio.Semaphore(CONCURRENCY_COUNT)

        # 创建所有任务
        tasks = []
        for idx, (question, row_num) in enumerate(zip(questions, row_numbers), start=1):
            task = asyncio.ensure_future(
                process_question(idx, row_num, total_questions, question, semaphore, headers)
            )
            tasks.append(task)

        # 使用 asyncio.wait 支持中断时保留已完成结果
        pending = set(tasks)
        completed_results = []  # 每个元素: (question_id, question, answer, hit_info, create_time, ask_time)

        # 增量保存计数
        completed_count = 0
        interrupted = False

        try:
            while pending:
                done, pending = await asyncio.wait(pending, return_when=asyncio.FIRST_COMPLETED)
                for task in done:
                    if not task.cancelled() and not task.exception():
                        completed_results.append(task.result())
                        completed_count += 1

                        # 增量保存
                        if INCREMENTAL_SAVE_INTERVAL > 0 and completed_count % INCREMENTAL_SAVE_INTERVAL == 0:
                            _save_results_incremental(wb_out, sheet_out, completed_results)
                            print(f"[增量保存] 已完成 {completed_count}/{total_questions} 题")

                # 处理异常任务
                for task in done:
                    if task.exception():
                        print(f"任务异常: {task.exception()}")

        except (asyncio.CancelledError, KeyboardInterrupt, EOFError):
            print("\n检测到中断，正在保存已完成的测试结果...")
            interrupted = True
            # 取消未完成的任务
            for task in tasks:
                if not task.done():
                    task.cancel()
            # 收集已完成任务结果
            for task in tasks:
                if task.done() and not task.cancelled() and not task.exception():
                    result = task.result()
                    if result not in completed_results:
                        completed_results.append(result)
                        completed_count += 1

        # 写入剩余结果到输出文件
        _write_results_to_sheet(sheet_out, completed_results)

        # 保存输出文件
        try:
            wb_out.save(output_file_name)
            wb_out.close()
            print(f"\n结果已保存到: {output_file_name}")
        except Exception as e:
            print(f"保存文件失败: {e}")

        # 统计
        success_count = sum(
            1 for r in completed_results
            if isinstance(r[2], str) and not r[2].startswith("[跳过]") and not r[2].startswith("Error")
        )
        error_count = sum(1 for r in completed_results if isinstance(r[2], str) and r[2].startswith("Error"))
        skip_count = sum(1 for r in completed_results if isinstance(r[2], str) and r[2].startswith("[跳过]"))

        # 输出统计信息
        if interrupted:
            print("测试被中断 - 已保存部分结果")
        else:
            print("测试完成！")
        print(f"总处理问题数: {len(tasks)} | 已完成: {len(completed_results)}")
        print(f"成功: {success_count} | 失败: {error_count} | 跳过: {skip_count}")

        stats = calculate_statistics()
        if stats:
            print("\n响应时间统计:")
            print(f"有效请求数: {stats['valid_requests']}")
            print(f"平均响应时间: {stats['avg_time']:.2f}秒")
            print(f"最短响应时间: {stats['min_time']:.2f}秒")
            print(f"最长响应时间: {stats['max_time']:.2f}秒")
            print(f"总耗时: {stats['total_time']:.2f}秒")

    except Exception as e:
        print(f"主程序执行失败: {e}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    # 检查依赖
    try:
        from Crypto.Cipher import PKCS1_v1_5
        from Crypto.PublicKey import RSA
    except ImportError:
        print("缺少必要的加密库, 请安装pycryptodome:")
        print("pip install pycryptodome")
        input("按回车键退出...")
        sys.exit(1)
    
    try:
        asyncio.run(main())
    except KeyboardInterrupt:
        print("\n用户中断执行")
    finally:
        input("\n按回车键退出...")